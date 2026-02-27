"""Маршруты для работы с объектами (адресами + камеры).

Этап B1 / Этап 2: «Метки/Объекты» как ядро сценария.

API:
  - /api/objects            CRUD
  - /api/objects/geo        облегчённый слой для карты
  - /api/objects/import     импорт CSV/XLSX
  - /api/objects/export/*   экспорт CSV/XLSX

Примечание:
  Проект использует Flask (не FastAPI). Поэтому здесь обычные Flask routes.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple

from compat_flask import jsonify, request, Response, current_app, session
from pydantic import BaseModel, ValidationError

from ..helpers import require_admin, parse_coord
from ..extensions import db
from ..models import Object, ObjectCamera
from ..security.rate_limit import check_rate_limit
from ..services.discovery_service import AutoDiscoveryService
from . import bp


# -------------------------
# Discover terminal request schema
# -------------------------


class DiscoverRequest(BaseModel):
    ip: str
    username: str
    password: str


# -------------------------
# Helpers
# -------------------------

_CAM_SPLIT = ';;'
_CAM_FIELD_SPLIT = '|'


def _rate_ident() -> str:
    """Identifier for rate limiting.

    Prefer admin username when in admin session; fall back to device header or IP.
    """
    if session.get('is_admin'):
        user = session.get('admin_username') or session.get('username') or 'admin'
        return f'admin:{user}'
    did = (request.headers.get('X-Device-ID') or request.headers.get('X-DEVICE-ID') or '').strip()
    if did:
        return f'dev:{did}'
    ip = (request.headers.get('CF-Connecting-IP') or request.remote_addr or 'ip')
    return f'ip:{ip}'


def _rate_limit_or_429(bucket: str, limit: int, window_seconds: int = 60):
    """Return a Flask response (429) if rate limited, else None."""
    try:
        ok, info = check_rate_limit(bucket=bucket, ident=_rate_ident(), limit=int(limit), window_seconds=int(window_seconds))
    except Exception:
        # If rate limiter is broken, do not block the request (best-effort).
        return None
    if ok:
        return None
    resp = jsonify({
        'error': 'rate_limited',
        'message': 'Too many requests',
        'bucket': bucket,
        **info.to_headers(),
    })
    resp.status_code = 429
    try:
        for k, v in info.http_headers().items():
            resp.headers[k] = v
    except Exception:
        pass
    return resp


# Заголовки импорта часто бывают на русском или с разными вариантами.
# Нормализуем ключи к каноническим.
_KEY_ALIASES = {
    # RU
    'адрес': 'name',
    'название': 'name',
    'имя': 'name',
    'описание': 'description',
    'примечание': 'description',
    'теги': 'tags',
    'тэги': 'tags',
    'категория': 'tags',
    'широта': 'lat',
    'долгота': 'lon',
    'долгота́': 'lon',
    'коорд': 'coords',
    'координаты': 'coords',
    'камеры': 'cameras',
    'камера': 'camera_url',
    'камера_url': 'camera_url',
    'url_камеры': 'camera_url',
    'название_камеры': 'camera_label',
    'тип_камеры': 'camera_type',
    # EN variants
    'address': 'name',
    'title': 'name',
    'notes': 'description',
    'category': 'tags',
    'lng': 'lon',
    'long': 'lon',
    'longitude': 'lon',
    'latitude': 'lat',
}


def _normalize_key(k: str) -> str:
    kk = (k or '').replace('\ufeff', '').strip().lower()
    kk = kk.replace(' ', '_')
    kk = _KEY_ALIASES.get(kk, kk)
    return kk


def _iso(dt: Optional[datetime]) -> Optional[str]:
    return dt.isoformat() if dt else None


def _parse_cameras_compact(value: str) -> List[Dict[str, Any]]:
    """Парсит компактную строку камер.

    Формат: "label|type|url;;label|type|url".
    Разрешаем неполные варианты:
      - "url" (одна камера без label/type)
      - "label|url" (если забыли type)
    """
    out: List[Dict[str, Any]] = []
    s = (value or '').strip()
    if not s:
        return out

    for chunk in [c for c in s.split(_CAM_SPLIT) if c.strip()]:
        parts = [p.strip() for p in chunk.split(_CAM_FIELD_SPLIT)]
        parts = [p for p in parts if p != '']
        label = None
        ctype = None
        url = None

        if len(parts) == 1:
            url = parts[0]
        elif len(parts) == 2:
            # предполагаем label|url
            label, url = parts
        else:
            label, ctype, url = parts[0], parts[1], parts[2]

        if url:
            out.append({
                'label': label or None,
                'type': ctype or None,
                'url': url,
            })
    return out


def _cameras_from_row(row: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Извлечь камеры из строки импорта.

    Поддержка:
      1) cameras="label|type|url;;..."
      2) camera_url + camera_label + camera_type
      3) camera1_url, camera2_url, ... (+ camera1_label, camera1_type)
    """
    # 1) compact
    compact = (row.get('cameras') or '').strip() if isinstance(row.get('cameras'), str) else row.get('cameras')
    if isinstance(compact, str) and compact.strip():
        return _parse_cameras_compact(compact)

    out: List[Dict[str, Any]] = []

    # 2) single camera
    cu = (row.get('camera_url') or '').strip() if isinstance(row.get('camera_url'), str) else ''
    if cu:
        out.append({
            'label': (row.get('camera_label') or '').strip() or None,
            'type': (row.get('camera_type') or '').strip() or None,
            'url': cu,
        })

    # 3) numbered cameras
    # camera1_url, camera2_url ...
    for k, v in list(row.items()):
        if not isinstance(k, str):
            continue
        m = re.match(r'^camera(\d+)_url$', k.strip().lower())
        if not m:
            continue
        idx = m.group(1)
        url = (v or '').strip() if isinstance(v, str) else ''
        if not url:
            continue
        label = (row.get(f'camera{idx}_label') or '').strip() or None
        ctype = (row.get(f'camera{idx}_type') or '').strip() or None
        out.append({'label': label, 'type': ctype, 'url': url})

    # если out пуст, вернём пустой список
    return out


def _object_to_export_row(obj: Object) -> Dict[str, Any]:
    cams = []
    for cam in (obj.cameras or []):
        label = cam.label or ''
        ctype = cam.type or ''
        url = cam.url or ''
        if not url:
            continue
        cams.append(f"{label}{_CAM_FIELD_SPLIT}{ctype}{_CAM_FIELD_SPLIT}{url}")
    cameras_str = _CAM_SPLIT.join(cams)

    return {
        'id': obj.id,
        'name': obj.name or '',
        'lat': obj.lat,
        'lon': obj.lon,
        'description': obj.description or '',
        'tags': obj.tags or '',
        'cameras': cameras_str,
        'created_at': _iso(obj.created_at),
        'updated_at': _iso(obj.updated_at),
    }


def _query_objects(q: str, tag: str):
    q = (q or '').strip().lower()
    tag = (tag or '').strip().lower()

    query = Object.query
    if q:
        like = f"%{q}%"
        query = query.filter((Object.name.ilike(like)) | (Object.description.ilike(like)))
    if tag:
        like = f"%{tag}%"
        query = query.filter(Object.tags.ilike(like))
    return query


@bp.post('/terminals/discover')
async def api_terminals_discover() -> Any:
    """Probe terminal credentials once and return discovery status."""
    require_admin("viewer")

    limited = _rate_limit_or_429('terminals_discover', limit=20, window_seconds=60)
    if limited is not None:
        return limited

    payload = request.get_json(silent=True) or {}
    try:
        data = DiscoverRequest.model_validate(payload)
    except ValidationError as exc:
        return jsonify({
            'status': 'error',
            'message': 'invalid_payload',
            'details': exc.errors(),
        }), 400

    result = await AutoDiscoveryService.probe_terminal(
        ip=data.ip,
        username=data.username,
        password=data.password,
    )

    status_code = 200 if result.get('status') == 'success' else 400
    return jsonify(result), status_code


# -------------------------
# Geo overlay
# -------------------------

@bp.get('/objects/geo')
def api_objects_geo() -> Any:
    """Лёгкий гео‑список объектов для слоя на карте."""
    require_admin("viewer")

    bbox = (request.args.get('bbox') or '').strip()
    west = south = east = north = None
    if bbox:
        try:
            parts = [float(x) for x in bbox.split(',')]
            if len(parts) == 4:
                west, south, east, north = parts
        except Exception:
            west = south = east = north = None

    q = (request.args.get('q') or '').strip().lower()
    tag = (request.args.get('tag') or '').strip().lower()
    try:
        limit = int(request.args.get('limit') or 1000)
    except Exception:
        limit = 1000
    limit = max(1, min(limit, 5000))

    query = Object.query.filter(Object.lat.isnot(None), Object.lon.isnot(None))
    if q:
        like = f"%{q}%"
        query = query.filter((Object.name.ilike(like)) | (Object.description.ilike(like)))
    if tag:
        like = f"%{tag}%"
        query = query.filter(Object.tags.ilike(like))

    if west is not None and south is not None and east is not None and north is not None:
        query = query.filter(Object.lon >= west, Object.lon <= east, Object.lat >= south, Object.lat <= north)

    objects = query.order_by(Object.created_at.desc()).limit(limit).all()

    out: List[Dict[str, Any]] = []
    for obj in objects:
        out.append({
            'id': obj.id,
            'lat': obj.lat,
            'lon': obj.lon,
            'name': obj.name,
            'tags': obj.tags,
            'camera_count': len(obj.cameras or []),
            'created_at': _iso(obj.created_at),
        })
    return jsonify(out), 200


# -------------------------
# CRUD
# -------------------------

@bp.get('/objects')
def list_objects() -> Any:
    """Вернуть список объектов (для admin UI)."""
    require_admin("viewer")
    q = request.args.get('q')
    tag = request.args.get('tag')
    # Ограничение, чтобы не вывалить мегабайты в браузер.
    try:
        limit = int(request.args.get('limit') or 500)
    except Exception:
        limit = 500
    limit = max(1, min(limit, 5000))

    lite = (request.args.get('lite') or '').strip().lower() in ('1', 'true', 'yes', 'on')

    query = _query_objects(q, tag).order_by(Object.created_at.desc()).limit(limit)
    objects = query.all()

    if lite:
        return jsonify([
            {
                'id': obj.id,
                'name': obj.name,
                'lat': obj.lat,
                'lon': obj.lon,
                'tags': obj.tags,
                'camera_count': len(obj.cameras or []),
                'created_at': _iso(obj.created_at),
                'updated_at': _iso(obj.updated_at),
            }
            for obj in objects
        ])

    return jsonify([obj.to_dict() for obj in objects])


@bp.post('/objects')
def create_object() -> Any:
    """Создать новый объект. Требует админа."""
    require_admin()
    rl = _rate_limit_or_429('objects_write', current_app.config.get('RATE_LIMIT_OBJECTS_WRITE_PER_MINUTE', 120), 60)
    if rl is not None:
        return rl
    data: Dict[str, Any] = request.get_json(silent=True) or {}
    name = (data.get('name') or data.get('address') or '').strip()
    description = (data.get('description') or data.get('notes') or '').strip()
    lat = parse_coord(data.get('lat'))
    lon = parse_coord(data.get('lon'))
    tags = (data.get('tags') or data.get('category') or '').strip()

    obj = Object(name=name, lat=lat, lon=lon, description=description, tags=tags)

    cams: List[Dict[str, Any]] = data.get('cameras') or []
    for cam in cams:
        url = (cam.get('url') or '').strip()
        if not url:
            continue
        label = (cam.get('label') or '').strip()
        ctype = (cam.get('type') or '').strip()
        obj.cameras.append(ObjectCamera(url=url, label=label or None, type=ctype or None))

    db.session.add(obj)
    db.session.commit()
    return jsonify({'id': obj.id}), 200


@bp.get('/objects/<int:object_id>')
def get_object(object_id: int) -> Any:
    require_admin("viewer")
    obj = Object.query.get_or_404(object_id)
    return jsonify(obj.to_dict())


@bp.put('/objects/<int:object_id>')
def update_object(object_id: int) -> Any:
    require_admin()
    rl = _rate_limit_or_429('objects_write', current_app.config.get('RATE_LIMIT_OBJECTS_WRITE_PER_MINUTE', 120), 60)
    if rl is not None:
        return rl
    obj: Optional[Object] = Object.query.get_or_404(object_id)
    data: Dict[str, Any] = request.get_json(silent=True) or {}

    name = data.get('name') or data.get('address')
    if name is not None:
        obj.name = str(name).strip()
    description = data.get('description') or data.get('notes')
    if description is not None:
        obj.description = str(description).strip()
    tags = data.get('tags') or data.get('category')
    if tags is not None:
        obj.tags = str(tags).strip()

    if 'lat' in data:
        obj.lat = parse_coord(data.get('lat'))
    if 'lon' in data:
        obj.lon = parse_coord(data.get('lon'))

    if 'cameras' in data:
        for cam in list(obj.cameras):
            db.session.delete(cam)
        cams: List[Dict[str, Any]] = data.get('cameras') or []
        for cam in cams:
            url = (cam.get('url') or '').strip()
            if not url:
                continue
            label = (cam.get('label') or '').strip()
            ctype = (cam.get('type') or '').strip()
            obj.cameras.append(ObjectCamera(url=url, label=label or None, type=ctype or None))

    db.session.commit()
    return jsonify({'id': obj.id}), 200


@bp.delete('/objects/<int:object_id>')
def delete_object(object_id: int) -> Any:
    require_admin("editor")
    rl = _rate_limit_or_429('objects_write', current_app.config.get('RATE_LIMIT_OBJECTS_WRITE_PER_MINUTE', 120), 60)
    if rl is not None:
        return rl
    obj: Optional[Object] = Object.query.get_or_404(object_id)
    db.session.delete(obj)
    db.session.commit()
    return jsonify({'ok': True, 'id': object_id}), 200


# -------------------------
# Export
# -------------------------

@bp.get('/objects/export/objects.csv')
def export_objects_csv() -> Response:
    require_admin("viewer")
    q = request.args.get('q')
    tag = request.args.get('tag')
    objects = _query_objects(q, tag).order_by(Object.created_at.desc()).all()

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=['id', 'name', 'lat', 'lon', 'description', 'tags', 'cameras', 'created_at', 'updated_at'],
        extrasaction='ignore'
    )
    writer.writeheader()
    for obj in objects:
        writer.writerow(_object_to_export_row(obj))

    csv_data = output.getvalue()
    return Response(
        csv_data,
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': 'attachment; filename="objects.csv"',
        },
    )


@bp.get('/objects/export/objects.xlsx')
def export_objects_xlsx() -> Response:
    require_admin("viewer")
    try:
        from openpyxl import Workbook
    except Exception as e:
        return jsonify({'error': 'openpyxl not installed', 'details': str(e)}), 500

    q = request.args.get('q')
    tag = request.args.get('tag')
    objects = _query_objects(q, tag).order_by(Object.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'objects'

    headers = ['id', 'name', 'lat', 'lon', 'description', 'tags', 'cameras', 'created_at', 'updated_at']
    ws.append(headers)

    for obj in objects:
        row = _object_to_export_row(obj)
        ws.append([row.get(h) for h in headers])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return Response(
        bio.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename="objects.xlsx"',
        },
    )




@bp.get('/objects/export/template.csv')
def export_objects_template_csv() -> Response:
    """Скачать шаблон CSV для импорта объектов."""
    require_admin("viewer")
    output = io.StringIO()
    headers = ['id', 'name', 'lat', 'lon', 'description', 'tags', 'cameras']
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction='ignore')
    writer.writeheader()
    writer.writerow({
        'id': '',
        'name': 'Пример: ул. Ленина, 1',
        'lat': 53.9000,
        'lon': 27.5667,
        'description': 'Описание объекта',
        'tags': 'видео;домофон',
        'cameras': 'Вход|video|https://example.com/stream1;;Двор|video|https://example.com/stream2',
    })
    return Response(
        output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': 'attachment; filename="objects_template.csv"'},
    )


@bp.get('/objects/export/template.xlsx')
def export_objects_template_xlsx() -> Response:
    """Скачать шаблон XLSX для импорта объектов."""
    require_admin("viewer")
    try:
        from openpyxl import Workbook
    except Exception as e:
        return jsonify({'error': 'openpyxl not installed', 'details': str(e)}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = 'objects_template'
    headers = ['id', 'name', 'lat', 'lon', 'description', 'tags', 'cameras']
    ws.append(headers)
    ws.append([
        '',
        'Пример: ул. Ленина, 1',
        53.9000,
        27.5667,
        'Описание объекта',
        'видео;домофон',
        'Вход|video|https://example.com/stream1;;Двор|video|https://example.com/stream2',
    ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        bio.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="objects_template.xlsx"'},
    )
# -------------------------
# Import
# -------------------------

@bp.post('/objects/import')
def import_objects() -> Any:
    """Импорт объектов из CSV/XLSX.

    multipart/form-data:
      file: *.csv или *.xlsx

    Логика:
      - если указан id и объект найден -> UPDATE
      - иначе -> CREATE

    Возвращает:
      { created, updated, errors_count, errors[] }
    """
    require_admin("editor")
    rl = _rate_limit_or_429('objects_import', current_app.config.get('RATE_LIMIT_OBJECTS_IMPORT_PER_MINUTE', 10), 60)
    if rl is not None:
        return rl

    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if not file or not file.filename:
        return jsonify({'error': 'No selected file'}), 400

    filename = (file.filename or '').lower()

    dry_run = (request.args.get('dry_run') or '').strip().lower() in ('1','true','yes','on')

    try:
        if filename.endswith('.xlsx'):
            rows = _read_xlsx_rows(file)
        else:
            # default: csv
            rows = _read_csv_rows(file)
    except Exception as e:
        return jsonify({'error': 'Failed to parse file', 'details': str(e)}), 400

    created = 0
    updated = 0
    errors: List[Dict[str, Any]] = []

    for idx, row in enumerate(rows, start=2):  # 1 is header
        try:
            res = _upsert_object_from_row(row)
            if res == 'created':
                created += 1
            elif res == 'updated':
                updated += 1
        except Exception as e:
            errors.append({'row': idx, 'error': str(e), 'data': row})

    if dry_run:
        # Откатываем изменения, но оставляем отчёт о том, что бы произошло.
        db.session.rollback()
    else:
        db.session.commit()

    return jsonify({
        'created': created,
        'updated': updated,
        'dry_run': dry_run,
        'errors_count': len(errors),
        'errors': errors[:200],
    }), 200


def _read_csv_rows(file) -> List[Dict[str, Any]]:
    raw = file.read()
    if isinstance(raw, bytes):
        # utf-8 with BOM support
        text = raw.decode('utf-8-sig', errors='replace')
    else:
        text = str(raw)

    sio = io.StringIO(text)
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=[',',';','	'])
    except Exception:
        dialect = csv.excel
    reader = csv.DictReader(sio, dialect=dialect)
    out: List[Dict[str, Any]] = []
    for r in reader:
        # normalize keys
        row: Dict[str, Any] = {}
        for k, v in (r or {}).items():
            if k is None:
                continue
            kk = _normalize_key(str(k))
            if not kk:
                continue
            row[kk] = (v.strip() if isinstance(v, str) else v)
        out.append(row)
    return out


def _read_xlsx_rows(file) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError(f'openpyxl not installed: {e}')

    data = file.read()
    bio = io.BytesIO(data)
    wb = load_workbook(bio, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []

    headers = [_normalize_key(str(h)) if h is not None else '' for h in header]

    out: List[Dict[str, Any]] = []
    for row_values in rows_iter:
        if not row_values:
            continue
        row: Dict[str, Any] = {}
        for i, val in enumerate(row_values):
            if i >= len(headers):
                continue
            key = headers[i]
            if not key:
                continue
            if isinstance(val, datetime):
                row[key] = val.isoformat()
            else:
                row[key] = val
        # skip fully empty
        if all((v is None or (isinstance(v, str) and not v.strip())) for v in row.values()):
            continue
        # normalize string values
        for k in list(row.keys()):
            v = row[k]
            if isinstance(v, str):
                row[k] = v.strip()
        out.append(row)
    return out


def _upsert_object_from_row(row: Dict[str, Any]) -> str:
    # Accept alternative keys
    obj_id_raw = row.get('id')
    obj_id = None
    try:
        if obj_id_raw is not None and str(obj_id_raw).strip() != '':
            obj_id = int(float(str(obj_id_raw).strip()))
    except Exception:
        obj_id = None

    name = (row.get('name') or row.get('address') or '').strip() if isinstance(row.get('name') or row.get('address') or '', str) else (row.get('name') or row.get('address') or '')
    description = row.get('description') or row.get('notes') or ''
    if not isinstance(description, str):
        description = '' if description is None else str(description)
    tags = row.get('tags') or row.get('category') or ''
    if not isinstance(tags, str):
        tags = '' if tags is None else str(tags)

    lat = parse_coord(row.get('lat'))
    lon = parse_coord(row.get('lon'))

    cams = _cameras_from_row(row)

    obj: Optional[Object] = None
    if obj_id is not None:
        obj = Object.query.get(obj_id)

    if obj:
        # UPDATE
        if name:
            obj.name = name
        if description is not None:
            obj.description = description
        if tags is not None:
            obj.tags = tags
        if 'lat' in row:
            obj.lat = lat
        if 'lon' in row:
            obj.lon = lon

        if cams:
            for cam in list(obj.cameras):
                db.session.delete(cam)
            for cam in cams:
                url = (cam.get('url') or '').strip()
                if not url:
                    continue
                obj.cameras.append(ObjectCamera(
                    url=url,
                    label=(cam.get('label') or None),
                    type=(cam.get('type') or None),
                ))

        db.session.add(obj)
        return 'updated'

    # CREATE
    if not name:
        raise ValueError('name/address required')

    obj = Object(name=name, lat=lat, lon=lon, description=description, tags=tags)
    for cam in cams:
        url = (cam.get('url') or '').strip()
        if not url:
            continue
        obj.cameras.append(ObjectCamera(
            url=url,
            label=(cam.get('label') or None),
            type=(cam.get('type') or None),
        ))

    db.session.add(obj)
    db.session.flush()
    return 'created'
