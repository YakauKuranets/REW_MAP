"""Маршруты аналитики.

Эндпоинты предоставляют агрегированную информацию о данных в системе.
Тяжёлая бизнес‑логика вынесена в :mod:`app.services.analytics_service`.
"""

import math
from datetime import datetime, timedelta
from flask import jsonify, make_response, request

from . import bp
from ..services.analytics_service import build_summary, build_audit_log, build_period_text


@bp.get('/summary')
def summary():
    """Вернуть сводную статистику по объектам и заявкам.

    Поддерживает параметры ``days`` (1–365) и ``zone_id`` для фильтрации.
    Если параметры не заданы, используются значения по умолчанию (7 дней,
    все зоны).
    """
    # Читаем параметры запроса. Используем строки и затем приводим к нужному
    # типу, чтобы корректно обрабатывать отсутствующие или пустые значения.
    days_param = request.args.get('days')
    zone_param = request.args.get('zone_id')
    try:
        days_val = int(days_param) if days_param else 7
    except (ValueError, TypeError):
        days_val = 7
    try:
        zone_val = int(zone_param) if zone_param else None
    except (ValueError, TypeError):
        zone_val = None
    data = build_summary(days_val, zone_val)
    return jsonify(data)


@bp.get('/text')
def text():
    """Вернуть облегчённую текстовую аналитику.

    Используется новым UI аналитики без графиков.
    Поддерживает параметр ``days`` (1–365). Фильтр по зонам
    сознательно не поддерживаем (для стабильности и простоты).
    """
    days_param = request.args.get('days')
    try:
        days_val = int(days_param) if days_param else 7
    except (ValueError, TypeError):
        days_val = 7

    data = build_period_text(days_val)
    return jsonify(data)


@bp.get('/summary.csv')
def summary_csv():
    """Вернуть сводную статистику в простом CSV‑виде.

    Поддерживает те же параметры ``days`` и ``zone_id``, что и JSON‑версии.
    Формат CSV предельно простой «ключ;значение», чтобы файл можно
    было легко разобрать внешними инструментами.
    """
    # Читаем параметры для фильтрации
    days_param = request.args.get('days')
    zone_param = request.args.get('zone_id')
    try:
        days_val = int(days_param) if days_param else 7
    except (ValueError, TypeError):
        days_val = 7
    try:
        zone_val = int(zone_param) if zone_param else None
    except (ValueError, TypeError):
        zone_val = None
    data = build_summary(days_val, zone_val)

    lines = []
    # Основные агрегаты
    lines.append(f"meta:total;{data.get('total', 0)}")
    lines.append(f"meta:pending;{data.get('pending', 0)}")
    lines.append(f"meta:approved;{data.get('approved', 0)}")
    lines.append(f"meta:rejected;{data.get('rejected', 0)}")
    # Метрика добавленных адресов за N дней
    added_n = data.get('added_last_n') or data.get('added_last_7d') or 0
    lines.append(f"meta:added_last_n;{added_n}")

    # Распределения по категориям / статусу доступа
    for key, value in (data.get('by_category') or {}).items():
        lines.append(f"by_category:{key};{value}")
    for key, value in (data.get('by_status') or {}).items():
        lines.append(f"by_status:{key};{value}")

    # Распределение по зонам
    for key, value in (data.get('by_zone') or {}).items():
        lines.append(f"by_zone:{key};{value}")

    # Распределение заявок по статусам
    for key, value in (data.get('pending_by_status') or {}).items():
        lines.append(f"pending_status:{key};{value}")

    # Таймлайн за выбранный период
    timeline = data.get('timeline_last_n') or data.get('timeline_last_7d') or []
    for row in timeline:
        date = row.get('date', '')
        addresses = row.get('addresses', 0)
        pending_created = row.get('pending_created', 0)
        approved = row.get('approved', 0)
        rejected = row.get('rejected', 0)
        lines.append(f"timeline:{date}:addresses;{addresses}")
        lines.append(f"timeline:{date}:pending_created;{pending_created}")
        lines.append(f"timeline:{date}:approved;{approved}")
        lines.append(f"timeline:{date}:rejected;{rejected}")

    csv_body = '\n'.join(lines)
    resp = make_response(csv_body)
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = 'attachment; filename=analytics_summary.csv'
    resp.headers['Cache-Control'] = 'public, max-age=60'
    return resp

@bp.get('/summary.xlsx')
def summary_xlsx():
    """Вернуть сводную статистику в формате Excel (XLSX).

    Поддерживает параметры ``days`` и ``zone_id``. Формирует таблицы по
    основным агрегатам и распределениям, что позволяет анализировать
    данные в Excel.
    """
    from openpyxl import Workbook
    from io import BytesIO

    # Читаем параметры для фильтрации
    days_param = request.args.get('days')
    zone_param = request.args.get('zone_id')
    try:
        days_val = int(days_param) if days_param else 7
    except (ValueError, TypeError):
        days_val = 7
    try:
        zone_val = int(zone_param) if zone_param else None
    except (ValueError, TypeError):
        zone_val = None
    data = build_summary(days_val, zone_val)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'

    # Основные метрики
    ws.append(['Метрика', 'Значение'])
    ws.append(['Всего адресов', data.get('total', 0)])
    ws.append(['Заявок в ожидании', data.get('pending', 0)])
    ws.append(['Одобрено', data.get('approved', 0)])
    ws.append(['Отклонено', data.get('rejected', 0)])
    # Метрика добавленных адресов за N дней
    added_n = data.get('added_last_n') or data.get('added_last_7d') or 0
    ws.append([f'Добавлено за {days_val} дней', added_n])

    # Распределения по категориям
    by_category = data.get('by_category') or {}
    ws_cat = wb.create_sheet(title='По категориям')
    ws_cat.append(['Категория', 'Количество'])
    for key, val in by_category.items():
        ws_cat.append([key, val])

    # Распределения по статусу доступа
    by_status = data.get('by_status') or {}
    ws_status = wb.create_sheet(title='По доступу')
    ws_status.append(['Статус доступа', 'Количество'])
    for key, val in by_status.items():
        ws_status.append([key, val])

    # Распределение по зонам
    by_zone = data.get('by_zone') or {}
    ws_zone = wb.create_sheet(title='По зонам')
    ws_zone.append(['Зона', 'Количество'])
    for key, val in by_zone.items():
        ws_zone.append([key, val])

    # Статусы заявок (pending)
    by_pending = data.get('pending_by_status') or {}
    ws_pending = wb.create_sheet(title='Статусы заявок')
    ws_pending.append(['Статус заявки', 'Количество'])
    for key, val in by_pending.items():
        ws_pending.append([key, val])

    # Таймлайн за выбранный период
    timeline = data.get('timeline_last_n') or data.get('timeline_last_7d') or []
    ws_timeline = wb.create_sheet(title=f'Таймлайн {days_val} дней')
    ws_timeline.append(['Дата', 'Добавлено адресов', 'Создано заявок', 'Одобрено', 'Отклонено'])
    for row in timeline:
        date = row.get('date', '')
        addresses = row.get('addresses', 0)
        pending_created = row.get('pending_created', 0)
        approved = row.get('approved', 0)
        rejected = row.get('rejected', 0)
        ws_timeline.append([date, addresses, pending_created, approved, rejected])

    # Сохраняем файл в буфер
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename=analytics_summary.xlsx'
    resp.headers['Cache-Control'] = 'public, max-age=60'
    return resp

@bp.get('/audit/recent')
def audit_recent():
    """Вернуть последние действия (аудит).

    Используется простая агрегированная лента событий по адресам
    и заявкам, без тяжёлых фильтров. Этого достаточно, чтобы
    супер‑админ мог быстро увидеть, что происходило в системе.
    """
    data = build_audit_log(limit=50)
    return jsonify(data)


@bp.get('/risk_heatmap')
def risk_heatmap():
    """
    AI Heatmaps: Генерация предиктивных зон риска.

    Анализирует исторические данные об инцидентах и инфраструктуре.
    Применяет функцию экспоненциального затухания по времени (Time Decay).
    Возвращает массив координат с весами вероятности для рендеринга WebGL.
    """
    try:
        # Импортируем модели локально, чтобы избежать циклических импортов
        from app.models import Address, PendingMarker

        # Анализируем данные за последние 30 дней
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)

        # Получаем недавние инциденты/заявки
        recent_pending = PendingMarker.query.filter(PendingMarker.created_at >= thirty_days_ago).all()
        # Получаем базовую инфраструктуру (для фона)
        addresses = Address.query.filter(Address.lat != None, Address.lon != None).limit(2000).all()

        heatmap_data = []

        # Математическая функция расчета веса (Time Decay)
        def calculate_weight(created_at, base_weight=0.5):
            if not created_at:
                return base_weight
            days_old = (datetime.utcnow() - created_at).days
            # Экспоненциальное затухание: чем старше событие, тем меньше его вес
            decay = math.exp(-days_old / 15.0)
            return round(base_weight + (0.5 * decay), 2)

        # 1. Загружаем инциденты (высокий приоритет)
        for p in recent_pending:
            if p.lat and p.lon:
                heatmap_data.append({
                    "lat": float(p.lat),
                    "lon": float(p.lon),
                    "weight": calculate_weight(p.created_at, base_weight=0.6),
                    "type": "incident"
                })

        # 2. Загружаем статические объекты (низкий фоновый приоритет)
        for a in addresses:
             if a.lat and a.lon:
                 heatmap_data.append({
                    "lat": float(a.lat),
                    "lon": float(a.lon),
                    "weight": 0.2,
                    "type": "infrastructure"
                 })

        return jsonify({
            "status": "success",
            "total_points": len(heatmap_data),
            "points": heatmap_data
        }), 200

    except Exception as e:
        import logging
        logging.getLogger("map-v12-analytics").error(f"AI Heatmap Error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500